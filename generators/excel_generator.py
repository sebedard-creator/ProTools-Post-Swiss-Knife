from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from utils.text_utils import transform_track_name_for_excel, extract_clip_name_and_comment, extract_track_code_and_actor

def create_excel(data, output_path):
    """Create an Excel spreadsheet from the parsed cue sheet data for ADR VALIDATION"""
    wb = Workbook()
    ws = wb.active
    ws.title = "ADR Validation"
    
    # Define styles
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=12, bold=True, color='2C5F8D')
    cell_font = Font(name='Arial', size=10)
    
    header_fill = PatternFill(start_color='E8F0F8', end_color='E8F0F8', fill_type='solid')
    comment_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')  # Soft yellow
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    current_row = 1
    
    # Add title (session name)
    ws.merge_cells(f'A{current_row}:E{current_row}')  # Changed to E for 5 columns
    title_cell = ws[f'A{current_row}']
    title_cell.value = data.get('title', data['session_name'])
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    current_row += 2
    
    # Process each track
    for track in data['tracks']:
        # Transform track name for Excel format and add line count
        display_track_name = transform_track_name_for_excel(track['name'])
        line_count = len(track['events'])
        lignes_label = "Ligne" if line_count == 1 else "Lignes"
        display_track_name = f"{display_track_name} - {line_count} {lignes_label}"
        
        # Add track name as section header
        ws.merge_cells(f'A{current_row}:E{current_row}')  # Changed to E for 5 columns
        header_cell = ws[f'A{current_row}']
        header_cell.value = display_track_name
        header_cell.font = header_font
        header_cell.alignment = left_alignment
        header_cell.fill = header_fill
        current_row += 1
        
        # Add events (no column headers, just data)
        for event in track['events']:
            # Extract clip name and comment
            clip_name, comment = extract_clip_name_and_comment(event['clip_name'])
            
            # Only include non-muted events OR include all (depending on requirements)
            # For now, including all events as in PDF
            ws[f'A{current_row}'] = event['start_time']
            ws[f'B{current_row}'] = event['end_time']
            ws[f'C{current_row}'] = event['duration']
            ws[f'D{current_row}'] = clip_name
            ws[f'E{current_row}'] = comment
            
            # Apply styling
            for col in ['A', 'B', 'C', 'D', 'E']:
                cell = ws[f'{col}{current_row}']
                cell.font = cell_font
                cell.alignment = left_alignment
                cell.border = thin_border
            
            # Apply yellow highlight to comment cell if it contains "VALIDER", "TBC", or "TBW"
            if comment and (('VALIDER' in comment.upper()) or ('TBC' in comment.upper()) or ('TBW' in comment.upper())):
                comment_cell = ws[f'E{current_row}']
                comment_cell.fill = comment_fill
            
            current_row += 1
        
        # Add spacing between tracks
        current_row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15  # Start Time
    ws.column_dimensions['B'].width = 15  # End Time
    ws.column_dimensions['C'].width = 12  # Duration
    ws.column_dimensions['D'].width = 40  # Clip Name
    ws.column_dimensions['E'].width = 35  # Comment
    
    # Save the workbook
    wb.save(output_path)

def create_excel_tc_order(data, output_path):
    """Create an Excel spreadsheet with all cues sorted by START TIME for ADR TC ORDER"""
    wb = Workbook()
    ws = wb.active
    ws.title = "ADR TC Order"

    # Define styles
    title_font = Font(name='Arial', size=14, bold=True)
    col_header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    cell_font = Font(name='Arial', size=10)
    cue_font = Font(name='Arial', size=10, bold=True)

    col_header_fill = PatternFill(start_color='2C5F8D', end_color='2C5F8D', fill_type='solid')
    comment_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')

    current_row = 1

    # Title
    ws.merge_cells(f'A{current_row}:H{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = data.get('title', data['session_name'])
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    current_row += 2

    # Column headers
    col_headers = ['CUE', 'PERSONNAGE', 'TC IN', 'TC OUT', 'DURÉE', 'TEXTE', 'COMMENTAIRE']
    col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    for col, header in zip(col_letters, col_headers):
        cell = ws[f'{col}{current_row}']
        cell.value = header
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    current_row += 1

    # Collect all events from all tracks, with per-character sequential counters
    all_events = []
    char_counters = {}  # track_code -> current count

    for track in data['tracks']:
        track_code, actor_name, _ = extract_track_code_and_actor(track['name'])
        display_name = transform_track_name_for_excel(track['name'])
        code_key = track_code if track_code else display_name

        # Count existing entries for this code to assign sequential numbers later
        if code_key not in char_counters:
            char_counters[code_key] = 0

        for event in track['events']:
            all_events.append({
                'track_code': track_code,
                'code_key': code_key,
                'display_name': display_name,
                'event': event
            })

    # Sort all events by start_time (string sort works because timecodes are zero-padded HH:MM:SS)
    all_events.sort(key=lambda x: x['event']['start_time'])

    # Assign sequential cue numbers per character (in TC order)
    char_seq = {}  # code_key -> running count
    for item in all_events:
        key = item['code_key']
        if key not in char_seq:
            char_seq[key] = 0
        char_seq[key] += 1
        item['seq'] = char_seq[key]

    # Write rows
    for item in all_events:
        event = item['event']
        clip_name, comment = extract_clip_name_and_comment(event['clip_name'])

        # Build cue label: VAL001 or 001 if no code
        if item['track_code']:
            cue_label = f"{item['track_code']}{item['seq']:03d}"
        else:
            cue_label = f"{item['seq']:03d}"

        ws[f'A{current_row}'] = cue_label
        ws[f'B{current_row}'] = item['display_name']
        ws[f'C{current_row}'] = event['start_time']
        ws[f'D{current_row}'] = event['end_time']
        ws[f'E{current_row}'] = event['duration']
        ws[f'F{current_row}'] = clip_name
        ws[f'G{current_row}'] = comment

        # Styling
        for col, align in zip(col_letters, [center_alignment, left_alignment, center_alignment,
                                             center_alignment, center_alignment, left_alignment, left_alignment]):
            cell = ws[f'{col}{current_row}']
            cell.font = cue_font if col == 'A' else cell_font
            cell.alignment = align
            cell.border = thin_border

        # Yellow highlight on comment if flagged
        if comment and (('VALIDER' in comment.upper()) or ('TBC' in comment.upper()) or ('TBW' in comment.upper())):
            ws[f'G{current_row}'].fill = comment_fill

        current_row += 1

    # Column widths
    ws.column_dimensions['A'].width = 10   # CUE
    ws.column_dimensions['B'].width = 30   # PERSONNAGE
    ws.column_dimensions['C'].width = 14   # TC IN
    ws.column_dimensions['D'].width = 14   # TC OUT
    ws.column_dimensions['E'].width = 12   # DURÉE
    ws.column_dimensions['F'].width = 40   # TEXTE
    ws.column_dimensions['G'].width = 30   # COMMENTAIRE

    wb.save(output_path)
