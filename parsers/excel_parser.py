import openpyxl

def parse_excel_markers(filepath):
    """
    Parses the Excel file to extract TC IN, Description, and Notes particulières.
    Dynamically finds the header row by looking for 'TC IN'.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    markers = []
    header_row_idx = -1
    col_tc_in = -1
    col_desc = -1
    col_notes = -1
    
    with open(filepath, 'rb') as f:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        
        # 1. Find the header row
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            for col_idx, cell_value in enumerate(row):
                if str(cell_value).strip().upper() == 'TC IN':
                    header_row_idx = row_idx
                    col_tc_in = col_idx
                    break
            
            if header_row_idx != -1:
                # We found 'TC IN'. Let's find 'Description' and 'Notes particulières'
                for col_idx, cell_value in enumerate(row):
                    val = str(cell_value).strip().lower()
                    if val == 'description':
                        col_desc = col_idx
                    elif 'notes particul' in val or val == 'notes':
                        col_notes = col_idx
                break
                
        if header_row_idx == -1:
            raise ValueError("Impossible de trouver la colonne 'TC IN' dans le fichier Excel.")
            
        # 2. Extract data
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if len(row) <= col_tc_in or row[col_tc_in] is None:
                continue # Skip empty rows
                
            tc_in = str(row[col_tc_in]).strip()
            if not tc_in:
                continue
            
            desc = ""
            if col_desc != -1 and len(row) > col_desc and row[col_desc]:
                desc = str(row[col_desc]).strip()
                
            notes = ""
            if col_notes != -1 and len(row) > col_notes and row[col_notes]:
                notes = str(row[col_notes]).strip()
                
            comment = desc
            if notes:
                if comment:
                    comment += f" - Note: {notes}"
                else:
                    comment = notes
                    
            markers.append({
                'tc_in': tc_in,
                'comment': comment
            })
            
        wb.close()
    return markers
